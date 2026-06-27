"""
demo.py

Demonstrates the enterprise data dictionary pipeline:
  1. Build a local SQLite database simulating a healthcare data warehouse
  2. Extract column-level metadata using an INFORMATION_SCHEMA-style query
  3. Generate field definitions (mock by default, live LLM if API key is set)
  4. Write results to a formatted Excel file matching the dictionary structure

Requirements:
    pip install openai openpyxl pandas

Usage:
    python demo.py                              # runs with mock definitions
    OPENAI_API_KEY=your_key python demo.py      # runs with live LLM
"""

import sqlite3
import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
    LLM_AVAILABLE = True
except ImportError:
    LLM_AVAILABLE = False


# --- Fictional healthcare data warehouse schema ---

SAMPLE_DDL = """
CREATE TABLE IF NOT EXISTS Dim_Patient (
    DWDimPatientID      INTEGER PRIMARY KEY,
    PatientID           INTEGER,
    FirstName           TEXT,
    LastName            TEXT,
    DateOfBirth         TEXT,
    Gender              TEXT,
    IsActive            INTEGER,
    EDW_FirstLoadDate   TEXT,
    EDW_LastUpdatedDate TEXT
);

CREATE TABLE IF NOT EXISTS Fact_Encounter (
    DWFactEncounterID   INTEGER PRIMARY KEY,
    PatientID           INTEGER,
    ProviderID          INTEGER,
    EncounterDate       TEXT,
    EncounterType       TEXT,
    DischargeStatus     TEXT,
    TotalCharges        REAL,
    EDW_FirstLoadDate   TEXT,
    EDW_LastUpdatedDate TEXT
);

CREATE TABLE IF NOT EXISTS Dim_Provider (
    DWDimProviderID     INTEGER PRIMARY KEY,
    ProviderID          INTEGER,
    ProviderName        TEXT,
    Specialty           TEXT,
    NPI                 TEXT,
    IsActive            INTEGER,
    EDW_FirstLoadDate   TEXT,
    EDW_LastUpdatedDate TEXT
);
"""

MOCK_DEFINITIONS = {
    "DWDimPatientID":      ("Surrogate key for patient dimension records", "", "confirmed"),
    "DWFactEncounterID":   ("Surrogate key for encounter fact records", "", "confirmed"),
    "DWDimProviderID":     ("Surrogate key for provider dimension records", "", "confirmed"),
    "PatientID":           ("Business identifier linking to the source system patient record", "", "confirmed"),
    "ProviderID":          ("Business identifier linking to the treating provider", "", "confirmed"),
    "FirstName":           ("Given name of the patient", "", "confirmed"),
    "LastName":            ("Family name of the patient", "", "confirmed"),
    "ProviderName":        ("Full name of the provider", "", "confirmed"),
    "DateOfBirth":         ("Date the patient was born", "Format: YYYY-MM-DD", "confirmed"),
    "Gender":              ("Gender identity recorded for the patient", "e.g. Male, Female, Unknown", "confirmed"),
    "Specialty":           ("Medical specialty associated with the provider", "e.g. Cardiology, Primary Care", "confirmed"),
    "NPI":                 ("National Provider Identifier assigned to the provider by CMS", "", "confirmed"),
    "IsActive":            ("Flag indicating whether the record is currently active", "e.g. 1 = active, 0 = inactive", "confirmed"),
    "EncounterDate":       ("Date the clinical encounter occurred", "Format: YYYY-MM-DD", "confirmed"),
    "EncounterType":       ("Category of the clinical encounter", "e.g. Inpatient, Outpatient, Emergency", "confirmed"),
    "DischargeStatus":     ("Status assigned to the patient at time of discharge", "e.g. Discharged, Transferred, Expired", "confirmed"),
    "TotalCharges":        ("Total billed charges associated with the encounter", "", "confirmed"),
    "EDW_FirstLoadDate":   ("Date the record was first loaded into the enterprise data warehouse", "", "confirmed"),
    "EDW_LastUpdatedDate": ("Most recent date the record was updated in the warehouse", "", "confirmed"),
}

PROMPT_TEMPLATE = """
You are building an enterprise data dictionary. Generate a concise definition for each
field listed below. Follow these rules:

- Use plain active voice, no terminal punctuation
- Surrogate keys (DWDim, DWFact prefix): define as "Surrogate key for..." or "Surrogate key linking to..."
- Business ID fields: define as "Business identifier linking to..."
- Boolean flags: include examples in Notes using format: e.g. 1 = active, 0 = inactive
- Confidence 90%+: mark as "confirmed". Otherwise: "review needed"

Return ONLY a JSON array. Each object must have: field, definition, notes, confidence.

Table: {table_name}
Fields:
{fields}
"""


def setup_db(path="demo_warehouse.db"):
    conn = sqlite3.connect(path)
    conn.executescript(SAMPLE_DDL)
    conn.commit()
    return conn


def extract_metadata(conn, table_name):
    rows = []
    for col in conn.execute(f"PRAGMA table_info({table_name})").fetchall():
        rows.append({
            "source_table": table_name,
            "field": col[1],
            "data_type": col[2] or "TEXT"
        })
    return rows


def define_mock(fields):
    results = []
    for f in fields:
        defn, notes, conf = MOCK_DEFINITIONS.get(
            f["field"], ("Definition pending review", "", "review needed")
        )
        results.append({**f, "definition": defn, "notes": notes, "confidence": conf})
    return results


def define_llm(fields, table_name, client):
    field_list = "\n".join(f["field"] for f in fields)
    prompt = PROMPT_TEMPLATE.format(table_name=table_name, fields=field_list)
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2
    )
    parsed = json.loads(response.choices[0].message.content.strip())
    results = []
    for item in parsed:
        match = next((f for f in fields if f["field"] == item["field"]), {})
        results.append({
            "source_table": match.get("source_table", table_name),
            "field": item["field"],
            "data_type": match.get("data_type", ""),
            "definition": item.get("definition", ""),
            "notes": item.get("notes", ""),
            "confidence": item.get("confidence", "review needed")
        })
    return results


def write_excel(records, output_path="sample_output.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"

    headers = ["Schema", "Source Table", "Field", "Data Type", "Definition", "Notes", "Confirmed"]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill = PatternFill("solid", fgColor="EEF2F7")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 20

    for row_num, record in enumerate(records, 2):
        values = [
            "Datawarehouse",
            record["source_table"],
            record["field"],
            record["data_type"],
            record["definition"],
            record["notes"],
            record["confidence"]
        ]
        fill = alt_fill if row_num % 2 == 0 else None
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = cell_font
            cell.border = border
            cell.alignment = left if col_num >= 5 else center
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_num].height = 30

    col_widths = [16, 22, 28, 14, 55, 35, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G1"

    wb.save(output_path)
    return output_path


def main():
    print("Setting up demo database...")
    conn = setup_db()

    tables = ["Dim_Patient", "Fact_Encounter", "Dim_Provider"]
    all_records = []

    api_key = os.environ.get("OPENAI_API_KEY")
    client = OpenAI(api_key=api_key) if (LLM_AVAILABLE and api_key) else None

    for table in tables:
        print(f"Processing {table}...")
        fields = extract_metadata(conn, table)
        if client:
            records = define_llm(fields, table, client)
        else:
            records = define_mock(fields)
        all_records.extend(records)

    output = write_excel(all_records)
    print(f"\nDone. {len(all_records)} fields written to {output}")
    conn.close()


if __name__ == "__main__":
    main()
